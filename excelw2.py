import openpyxl
from openpyxl.chart import BarChart, Reference
wb = openpyxl.load_workbook('Book1.xlsx')
sheet=wb['Sheet1']

for row in range(2, sheet.max_row+1):
    excell=sheet.cell(row,4)
    corrected_price = excell.value*1.1
    corrected_price_cell=sheet.cell(row,5)
    corrected_price_cell.value=corrected_price
value=Reference(sheet, min_row=1, max_row=sheet.max_row, min_col=4, max_col=5)
chart=BarChart()
chart.add_data(value)
sheet.add_chart(chart, 'E2')
wb.save('python1.xlsx')